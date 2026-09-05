"""Baut eine Review-Tabelle fuer den vorgeschlagenen Wochen-Dienstplan (Auftrag
t.duong 05.09.2026) aus dienstplan_woche_vorschlag.json.
Aufruf: python build_dienstplan_excel.py
Ausgabe: dienstplan_woche_vorschlag.xlsx im output/-Ordner.
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output')

with open(os.path.join(OUT_DIR, 'dienstplan_woche_vorschlag.json'), 'r', encoding='utf-8') as f:
    data = json.load(f)

ergebnisse = data['ergebnisse']
uebersprungen = data['uebersprungen']
woche_start = data['wocheStart']

FONT = 'Arial'
HEADER_FILL = PatternFill(start_color='1B2430', end_color='1B2430', fill_type='solid')
HEADER_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
NOTE_FONT = Font(name=FONT, italic=True, size=9, color='6B7481')
TITLE_FONT = Font(name=FONT, bold=True, size=14)
THIN = Side(style='thin', color='DDE1E7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
VOLL_FILL = PatternFill(start_color='FCEFDA', end_color='FCEFDA', fill_type='solid')
FREI_FILL = PatternFill(start_color='F2F1EE', end_color='F2F1EE', fill_type='solid')
WARN_FILL = PatternFill(start_color='FBEAEA', end_color='FBEAEA', fill_type='solid')
OK_FONT = Font(name=FONT, size=10, bold=True, color='1E8449')

TAGE = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa']
VOLLE_TAGE = {'Mo', 'Do', 'Fr', 'Sa'}

wb = Workbook()
ws = wb.active
ws.title = 'Dienstplan-Vorschlag'
ws['A1'] = f'Dienstplan-Vorschlag — Woche ab {woche_start} (Mo)'
ws['A1'].font = TITLE_FONT
ws['A2'] = ('Schichtlängen an Basis der Vertragsstunden (Welo sollStd) berechnet, Startzeiten aus der historisch '
            'häufigsten Schicht je Wochentag übernommen. Orange hinterlegte Tage (Mo/Do/Fr/Sa) haben mehr Stunden '
            '("volle" Tage laut t.duong), Di/Mi sind ruhigere Tage. Wochensumme ist immer ≥ Vertragsstunden.')
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:K2')
ws['A2'].alignment = Alignment(wrap_text=True)
ws.row_dimensions[2].height = 30

headers = ['PersonalNr', 'Name', 'Filiale', 'Soll-Std/Wo.'] + TAGE + ['Wochensumme']
header_row = 4
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal='center')

ergebnisse_sorted = sorted(ergebnisse, key=lambda e: (e.get('filiale', ''), e.get('name', '')))
row = header_row + 1
for e in ergebnisse_sorted:
    vals = [e['id'], e['name'], e['filiale'], e['sollStd']] + [e['tage'][t] for t in TAGE] + [e['wochensumme']]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        if i == len(vals):
            c.font = OK_FONT
            c.alignment = Alignment(horizontal='center')
        if 5 <= i <= 10:
            tag = TAGE[i - 5]
            if vals[i - 1] == 'Frei':
                c.fill = FREI_FILL
            elif tag in VOLLE_TAGE:
                c.fill = VOLL_FILL
            c.alignment = Alignment(horizontal='center')
    row += 1

widths = [11, 26, 44, 11, 13, 13, 13, 13, 13, 13, 13]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = f'A{header_row + 1}'

ws2 = wb.create_sheet('Übersprungen (manuell prüfen)')
ws2['A1'] = 'Nicht automatisch geplant — keine Schicht-Historie in Welo'
ws2['A1'].font = TITLE_FONT
ws2['A2'] = ('Diese Personen haben keinen erkennbaren Schicht-Rhythmus (Gebietsleiter-Rollen ohne feste Filiale, '
             'oder Datenlücke) — bitte manuell im Dienstplan eintragen.')
ws2['A2'].font = NOTE_FONT
ws2.merge_cells('A2:D2')
h2 = ['PersonalNr', 'Name', 'Filiale', 'Grund']
for i, h in enumerate(h2, start=1):
    c = ws2.cell(row=4, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
row = 5
for u in uebersprungen:
    vals = [u['id'], u['name'], u.get('filiale', ''), u['grund']]
    for i, v in enumerate(vals, start=1):
        c = ws2.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        c.fill = WARN_FILL
    row += 1
widths2 = [11, 30, 44, 40]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

out_path = os.path.join(OUT_DIR, 'dienstplan_woche_vorschlag.xlsx')
wb.save(out_path)
print('Gespeichert:', out_path)
print('Geplant:', len(ergebnisse), '| Übersprungen:', len(uebersprungen))
