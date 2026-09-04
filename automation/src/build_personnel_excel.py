"""Baut personal_und_filialen.xlsx aus branch_personnel_export.json (siehe
export-personnel-data.js). Aufruf: python build_personnel_excel.py
Ausgabe: personal_und_filialen.xlsx im selben Ordner.
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output')

with open(os.path.join(OUT_DIR, 'branch_personnel_export.json'), 'r', encoding='utf-8') as f:
    data = json.load(f)

filialen = data['filialen']
emps = data['emps']

FONT = 'Arial'
HEADER_FILL = PatternFill(start_color='1B2430', end_color='1B2430', fill_type='solid')
HEADER_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
NOTE_FONT = Font(name=FONT, italic=True, size=9, color='6B7481')
TITLE_FONT = Font(name=FONT, bold=True, size=14)
THIN = Side(style='thin', color='DDE1E7')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WEST_FILL = PatternFill(start_color='E4F1EE', end_color='E4F1EE', fill_type='solid')
OST_FILL = PatternFill(start_color='FCEFDA', end_color='FCEFDA', fill_type='solid')
WARN_FILL = PatternFill(start_color='FBEAEA', end_color='FBEAEA', fill_type='solid')

wb = Workbook()

ws = wb.active
ws.title = 'Filialen'
ws['A1'] = 'Filialen — Adressen (Master-Liste)'
ws['A1'].font = TITLE_FONT
ws['A2'] = 'Neue Filiale? Einfach eine Zeile unten anhängen (MarktNr, Name, Ort, Straße, PLZ, Region) — Claude berechnet daraus die nächstgelegene Filiale für die Zweitfiliale-Zuordnung.'
ws['A2'].font = NOTE_FONT
ws.merge_cells('A2:G2')

headers = ['MarktNr', 'Filiale (voller Name)', 'Ort', 'Straße', 'PLZ', 'Region', 'Aktiv']
header_row = 4
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

filialen_sorted = sorted(filialen, key=lambda f: (f.get('region', ''), f.get('ort', '')))
row = header_row + 1
for f in filialen_sorted:
    vals = [f.get('marktNr', ''), f.get('name', ''), f.get('ort', ''), f.get('strasse', ''), f.get('plz', ''), f.get('region', ''), 'Ja' if f.get('active') else 'Nein']
    fill = WEST_FILL if f.get('region') == 'west' else (OST_FILL if f.get('region') == 'ost' else None)
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        if fill:
            c.fill = fill
    row += 1

widths = [10, 46, 20, 28, 8, 9, 7]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = f'A{header_row + 1}'

ws2 = wb.create_sheet('Mitarbeiter')
ws2['A1'] = 'Mitarbeiter — Filiale, Zweitfiliale & Wohnadresse'
ws2['A1'].font = TITLE_FONT
ws2['A2'] = 'Wohnadresse aus Welo (Personal > Suche "*" > Personal-Nr. anklicken, Felder Strasse/PLZ/Ort). Rot markiert = noch keine Zweitfiliale (meist weil keine Filiale in zumutbarer Nähe liegt).'
ws2['A2'].font = NOTE_FONT
ws2.merge_cells('A2:H2')

headers2 = ['PersonalNr', 'Name', 'Stammfiliale', 'Zweitfiliale(n)', 'Wohnort (PLZ, Ort)', 'Straße', 'Region', 'Shopleiter']
header_row2 = 4
for i, h in enumerate(headers2, start=1):
    c = ws2.cell(row=header_row2, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

emps_sorted = sorted(emps, key=lambda e: (e.get('region', ''), e.get('filiale', ''), e.get('name', '')))
row = header_row2 + 1
for e in emps_sorted:
    zweit = ', '.join(e.get('zweit') or [])
    wohnort = f"{e.get('plz','')} {e.get('ort','')}".strip()
    vals = [e.get('id', ''), e.get('name', ''), e.get('filiale', ''), zweit, wohnort, e.get('strasse', ''), e.get('region', ''), 'Ja' if e.get('shopleiter') else '']
    fill = WEST_FILL if e.get('region') == 'west' else (OST_FILL if e.get('region') == 'ost' else None)
    if not zweit:
        fill = WARN_FILL
    for i, v in enumerate(vals, start=1):
        c = ws2.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        if fill:
            c.fill = fill
    row += 1

widths2 = [11, 26, 46, 46, 20, 26, 9, 11]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = f'A{header_row2 + 1}'

out_path = os.path.join(OUT_DIR, 'personal_und_filialen.xlsx')
wb.save(out_path)
print('Gespeichert:', out_path)
print('Filialen:', len(filialen), '| Mitarbeiter:', len(emps))
